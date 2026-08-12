"""Excel workbook — partner debate surface quality gates."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.excel import CORE_TABS, APP_TABS, PIPELINE_HEADERS, build_workbook


SAMPLE_COMPANIES = [
    {
        "id": "c_a",
        "slug": "alpha",
        "name": "Alpha",
        "one_liner": "Eval infra for frontier labs",
        "sector_theme": "AI Infrastructure & Compute Stack",
        "theme_id": "ai_infra",
        "stage": "Series B",
        "pipeline_bucket": "dominant_tech_growth",
        "last_round_size_m": 40,
        "last_round_date": "2026-06-01",
        "valuation_est_m": 400,
        "lead_investor": "a16z",
        "investors": ["a16z", "Sequoia Capital"],
        "tier1_count": 2,
        "headcount": 80,
        "headcount_6m_growth_pct": 30,
        "last_signal_date": "2026-08-01",
        "thesis_score": 88,
        "recommendation": "Deep Dive",
        "why_now": "Tier-1 density + eval demand.",
        "relative_rank": "#1 / 6 in AI infra × Series B",
        "commentary_summary": "Labs adopting weekly.",
        "sources": ["seed"],
        "is_stale": False,
        "score_breakdown": {
            "thesis_fit": 90,
            "team_quality": 88,
            "cap_table": 92,
            "traction": 85,
            "moat": 80,
            "valuation": 70,
            "runway": 75,
            "tam_exit": 78,
            "timing": 90,
        },
    },
    {
        "id": "c_b",
        "slug": "beta",
        "name": "Beta",
        "one_liner": "Quiet seed cyber name",
        "sector_theme": "Cybersecurity",
        "theme_id": "cyber",
        "stage": "Seed",
        "pipeline_bucket": "dominant_tech_growth",
        "last_round_size_m": 4,
        "last_round_date": "2025-11-01",
        "valuation_est_m": 40,
        "lead_investor": "Initialized Capital",
        "investors": ["Initialized Capital"],
        "tier1_count": 0,
        "headcount": 12,
        "headcount_6m_growth_pct": 10,
        "last_signal_date": "2025-12-01",
        "thesis_score": 62,
        "recommendation": "Watch",
        "why_now": "Early; wait for T1.",
        "relative_rank": "#4 / 5 in cyber × Seed",
        "sources": ["seed"],
        "is_stale": True,
        "review_status": "Pending Partner Review",
        "score_breakdown": {
            "thesis_fit": 70,
            "team_quality": 60,
            "cap_table": 40,
            "traction": 50,
            "moat": 55,
            "valuation": 80,
            "runway": 60,
            "tam_exit": 65,
            "timing": 50,
        },
    },
]

SAMPLE_COMMENTARY = [
    {
        "company_id": "c_a",
        "company_name": "Alpha",
        "source": "Hacker News",
        "quote_or_summary": "Engineers calling this the eval layer that stuck.",
        "sentiment": "positive",
        "credibility_tier": "high",
        "captured_at": "2026-08-02",
    }
]

SAMPLE_NEWS = [
    {
        "title": "Agent identity is the next Okta",
        "source": "The Generalist",
        "url": "https://example.com/agent",
        "published_at": "2026-08-04",
        "why_it_matters": "Validates cyber × AI auth theme.",
        "related_themes": ["Cybersecurity"],
    }
]

SAMPLE_PEERS = [
    {
        "firm": "a16z",
        "company_name": "Alpha",
        "round": "Series B",
        "date": "2026-06-01",
        "theme": "AI Infrastructure & Compute Stack",
        "on_thesis_flag": True,
        "thesis_shift": False,
        "notes": "Led round",
    },
    {
        "firm": "Tiger Global",
        "company_name": "Alpha",
        "round": "Series B",
        "date": "2026-06-01",
        "theme": "Defence Tech",
        "on_thesis_flag": False,
        "thesis_shift": True,
        "notes": "Off-thesis for Tiger",
    },
]

SAMPLE_SECTORS = [
    {
        "subsector": "RL environment factories",
        "parent_theme": "AI Infrastructure & Compute Stack",
        "heat_score": 90,
        "consensus_level": "Contrarian",
        "evidence": ["arXiv velocity"],
        "top_companies": ["Alpha"],
        "why_thirdbase_cares": "Upstream of model capability",
    }
]


def test_build_workbook_core_and_app_tabs(tmp_path: Path):
    out = tmp_path / "Thirdbase_Deal_Pipeline.xlsx"
    path = build_workbook(
        SAMPLE_COMPANIES,
        SAMPLE_COMMENTARY,
        SAMPLE_NEWS,
        SAMPLE_PEERS,
        SAMPLE_SECTORS,
        meta={
            "last_refreshed": "2026-08-12T00:00:00+00:00",
            "app_base": "http://localhost:3000",
            "alerts": [
                {
                    "severity": "high",
                    "alert_type": "tier1_density",
                    "title": "2+ Tier-1 on Alpha",
                    "body": "Immediate routing",
                    "company_id": "c_a",
                    "created_at": "2026-08-12",
                    "brief_url": "/company/alpha",
                }
            ],
            "digest": {
                "subject": "Thirdbase Signal · Test digest",
                "generated_at": "2026-08-12",
                "deals": [
                    {
                        "name": "Alpha",
                        "recommendation": "Deep Dive",
                        "score": 88,
                        "sector": "AI Infrastructure",
                        "stage": "Series B",
                        "why_now": "Eval layer",
                        "slug": "alpha",
                    }
                ],
                "sector_calls": [],
                "news": [],
            },
            "golden_insights": [
                {
                    "urgency": "now",
                    "kind": "alpha",
                    "title": "Call Alpha",
                    "insight": "Quiet cap table",
                    "action": "Partner takes the call",
                    "score": 95,
                }
            ],
            "golden_brief": {"must_do": ["Call Alpha"], "proprietary": ["Alpha"]},
            "judgment": {
                "mix": {"dominant_pct": 100, "tactical_pct": 0, "status": "hard_drift", "counsel": "Allow tactical"},
                "misses": [],
                "founder_radar": [],
                "freshness": [],
            },
            "thesis_shifts": [
                {
                    "firm": "Tiger Global",
                    "company_name": "Alpha",
                    "theme": "Defence Tech",
                    "date": "2026-06-01",
                    "notes": "Off-thesis",
                    "thesis_shift": True,
                }
            ],
            "peer_firms": [],
            "heatmap": [],
            "peer_matrix": {},
        },
        out_path=out,
    )
    assert path.exists()
    wb = load_workbook(path)
    names = wb.sheetnames

    assert names[0] == "Index"
    for tab in CORE_TABS:
        assert tab in names, tab
    for tab in APP_TABS:
        assert tab in names, tab

    # No fluff cover marketing sheet
    assert "Cover" not in names
    assert "Golden Insights" not in names  # renamed Actions

    pipe = wb["Pipeline"]
    headers = [c.value for c in pipe[1]]
    for required in [
        "Company",
        "One-line description",
        "Thesis score",
        "Recommendation",
        "Link to full brief",
        "Investor commentary summary",
        "Why now",
    ]:
        assert required in headers
    assert headers == PIPELINE_HEADERS
    assert pipe.cell(2, 1).value == "Alpha"
    link = pipe.cell(2, PIPELINE_HEADERS.index("Link to full brief") + 1)
    assert link.hyperlink is not None
    assert "alpha" in str(link.hyperlink.target or link.hyperlink)

    hot = wb["Hot Deals"]
    assert hot.max_row >= 2
    assert hot.cell(2, 1).value == "Alpha"

    stale = wb["Stale"]
    assert stale.cell(2, 1).value == "Beta"
    assert stale.cell(2, 5).value is not None  # days silent

    alerts = wb["Alerts"]
    assert alerts.cell(2, 1).value == "high"
    assert "Alpha" in str(alerts.cell(2, 3).value)

    digest = wb["Digest"]
    assert "Alpha" in str(digest.cell(5, 1).value)

    scores = wb["Scores"]
    assert scores.cell(2, 1).value == "Alpha"
    assert scores.cell(2, 3).value == 88

    actions = wb["Actions"]
    assert actions.cell(6, 1).value == "now"

    index = wb["Index"]
    assert "Deal Pipeline" in str(index["A1"].value)
    assert index["A21"].value == "Full refresh"


def test_pipeline_headers_match_partner_brief():
    # Brief-required columns remain present (extra debate cols allowed after)
    required = [
        "Company",
        "One-line description",
        "Sector",
        "Stage",
        "Last round size ($M)",
        "Last round date",
        "Valuation ($M)",
        "Lead investor",
        "Tier 1 count",
        "Headcount",
        "6-month growth (%)",
        "Thesis score",
        "Theme tag",
        "Last signal date",
        "Recommendation",
        "Link to full brief",
        "Investor commentary summary",
    ]
    for h in required:
        assert h in PIPELINE_HEADERS
