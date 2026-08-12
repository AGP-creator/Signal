"""Thirdbase Signal — living Excel workbook (partner debate surface).

Design rules:
- Extreme usefulness, zero marketing fluff
- One job per tab; filterable tables with freeze + AutoFilter
- Same tabs as /workbook; downloadable via /api/workbook
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.ingest.discovery import classify_news_kind

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUT = ROOT / "data" / "output" / "Thirdbase_Deal_Pipeline.xlsx"
DEFAULT_APP_BASE = "http://localhost:3000"

# Visual system — navy headers, traffic-light recs, no decorative chrome
HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1B2A4A")
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="1B2A4A")
BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(color="0563C1", underline="single", name="Calibri", size=10)
DEEP_FILL = PatternFill("solid", fgColor="C6EFCE")
WATCH_FILL = PatternFill("solid", fgColor="FFEB9C")
PASS_FILL = PatternFill("solid", fgColor="D9D9D9")
HIGH_FILL = PatternFill("solid", fgColor="FFC7CE")
MED_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Brief-ordered core tabs (FEATURE_GUIDE / partner ask)
CORE_TABS = [
    "Pipeline",
    "Hot Deals",
    "Watchlist",
    "Sector of Tomorrow",
    "Peer Set Activity",
    "Co-investor Heatmap",
    "News Worth Reading",
    "Investor Commentary",
    "Stale",
]

# Signal app extracts — same store as web desks, table form
APP_TABS = [
    "Alerts",
    "Digest",
    "Scores",
    "Thesis Shifts",
    "Peer Firm Dossiers",
    "Actions",
    "Judgment",
]

PIPELINE_HEADERS = [
    "Company",
    "One-line description",
    "Sector",
    "Stage",
    "Bucket",
    "Last round size ($M)",
    "Last round date",
    "Valuation ($M)",
    "Lead investor",
    "Tier 1 count",
    "Headcount",
    "6-month growth (%)",
    "Thesis score",
    "Theme tag",
    "Last signal date",
    "Recommendation",
    "Link to full brief",
    "Investor commentary summary",
    "Why now",
    "Relative rank",
]

SCORE_DIMS = [
    "thesis_fit",
    "team_quality",
    "cap_table",
    "traction",
    "moat",
    "valuation",
    "runway",
    "tam_exit",
    "timing",
]

SCORE_DIM_LABELS = {
    "thesis_fit": "Thesis fit",
    "team_quality": "Team",
    "cap_table": "Cap table",
    "traction": "Traction",
    "moat": "Moat",
    "valuation": "Valuation",
    "runway": "Runway",
    "tam_exit": "TAM/exit",
    "timing": "Timing",
}

# Tab accent colors (Excel tabColor ARGB without alpha prefix)
TAB_COLORS = {
    "Index": "1B2A4A",
    "Pipeline": "1B2A4A",
    "Hot Deals": "C00000",
    "Watchlist": "BF8F00",
    "Sector of Tomorrow": "2E75B6",
    "Peer Set Activity": "548235",
    "Co-investor Heatmap": "548235",
    "News Worth Reading": "7030A0",
    "Investor Commentary": "7030A0",
    "Stale": "833C0C",
    "Alerts": "C00000",
    "Digest": "C00000",
    "Scores": "1B2A4A",
    "Thesis Shifts": "548235",
    "Peer Firm Dossiers": "548235",
    "Actions": "C00000",
    "Judgment": "833C0C",
}


def _clip(val: Any, n: int = 160) -> str:
    s = str(val or "").strip()
    if len(s) <= n:
        return s
    return s[: n - 1].rstrip() + "…"


def _days_since(iso: Any, as_of: Optional[date] = None) -> Optional[int]:
    if not iso:
        return None
    as_of = as_of or date.today()
    try:
        d = datetime.strptime(str(iso)[:10], "%Y-%m-%d").date()
        return (as_of - d).days
    except Exception:
        return None


def _brief_url(company: dict[str, Any], app_base: str) -> str:
    slug = company.get("slug") or company.get("id") or company.get("brief_id") or ""
    return f"{app_base.rstrip('/')}/company/{slug}"


def _set_tab_color(ws, name: str) -> None:
    color = TAB_COLORS.get(name)
    if color:
        ws.sheet_properties.tabColor = color


def _style_header(ws, ncols: int, row: int = 1) -> None:
    ws.row_dimensions[row].height = 28
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(ncols)}{row}"


def _autosize(ws, widths: Optional[dict[int, int]] = None, max_width: int = 42) -> None:
    """Size columns from sample content; optional fixed overrides by 1-based col index."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        idx = col[0].column
        if widths and idx in widths:
            ws.column_dimensions[letter].width = widths[idx]
            continue
        length = 0
        for cell in col[:50]:
            length = max(length, min(len(str(cell.value or "")), max_width))
        ws.column_dimensions[letter].width = max(10, length + 2)


def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    start_row: int = 1,
    widths: Optional[dict[int, int]] = None,
    number_cols: Optional[dict[int, str]] = None,
) -> int:
    """Write a filterable table. Returns next free row."""
    for i, h in enumerate(headers, 1):
        ws.cell(start_row, i, h)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if number_cols and c_idx in number_cols and isinstance(val, (int, float)):
                cell.number_format = number_cols[c_idx]
    _style_header(ws, len(headers), row=start_row)
    if rows:
        ws.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
        )
    _autosize(ws, widths=widths)
    return start_row + len(rows) + 2


def _linkify(ws, col: int, start_row: int, n: int, label: str = "Open brief") -> None:
    for r_idx in range(start_row, start_row + n):
        cell = ws.cell(r_idx, col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.value = label
            cell.font = LINK_FONT


def _rec_formatting(ws, col: int, n_rows: int) -> None:
    if n_rows <= 0:
        return
    letter = get_column_letter(col)
    rng = f"{letter}2:{letter}{n_rows + 1}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Deep Dive"'], fill=DEEP_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Watch"'], fill=WATCH_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL)
    )
    dv = DataValidation(type="list", formula1='"Deep Dive,Watch,Pass"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def _score_scale(ws, col: int, n_rows: int) -> None:
    if n_rows <= 0:
        return
    letter = get_column_letter(col)
    ws.conditional_formatting.add(
        f"{letter}2:{letter}{n_rows + 1}",
        ColorScaleRule(
            start_type="num",
            start_value=50,
            start_color="F8696B",
            mid_type="num",
            mid_value=75,
            mid_color="FFEB84",
            end_type="num",
            end_value=95,
            end_color="63BE7B",
        ),
    )


def _coinvestor_pairs(companies: list[dict[str, Any]]) -> list[list[Any]]:
    pair_data: dict[tuple[str, str], dict[str, Any]] = {}
    for c in companies:
        invs = sorted(set(c.get("investors") or []))
        for i in range(len(invs)):
            for j in range(i + 1, len(invs)):
                a, b = invs[i], invs[j]
                key = (a, b)
                rec = pair_data.setdefault(
                    key,
                    {
                        "count": 0,
                        "themes": set(),
                        "last": c.get("last_round_date") or "",
                        "deal": c.get("name"),
                    },
                )
                rec["count"] += 1
                rec["themes"].add(c.get("sector_theme") or "")
                if (c.get("last_round_date") or "") >= rec["last"]:
                    rec["last"] = c.get("last_round_date") or ""
                    rec["deal"] = c.get("name")
    rows = []
    for (a, b), rec in sorted(pair_data.items(), key=lambda x: -x[1]["count"]):
        rows.append(
            [
                a,
                b,
                rec["count"],
                ", ".join(sorted(t for t in rec["themes"] if t)),
                rec["deal"],
                rec["last"],
            ]
        )
    return rows[:80]


def _write_peer_matrix(ws, matrix: dict[str, Any], start_row: int = 1) -> int:
    firms = matrix.get("firms") or []
    companies = matrix.get("companies") or []
    cells = matrix.get("cells") or []
    if not firms or not companies:
        return start_row

    lookup = {(c.get("firm_slug"), c.get("company_id")): c for c in cells}

    ws.cell(start_row, 1, "Investor x company (Y = invested, L = lead)")
    ws.cell(start_row, 1).font = LABEL_FONT
    header_row = start_row + 1
    ws.cell(header_row, 1, "Firm")
    for i, co in enumerate(companies, 2):
        ws.cell(header_row, i, co.get("name"))
    _style_header(ws, len(companies) + 1, row=header_row)

    for r_idx, firm in enumerate(firms, header_row + 1):
        ws.cell(r_idx, 1, firm.get("name"))
        for c_idx, co in enumerate(companies, 2):
            hit = lookup.get((firm.get("slug"), co.get("id")))
            if not hit:
                val = ""
            elif hit.get("is_lead"):
                val = "L"
            else:
                val = "Y"
            cell = ws.cell(r_idx, c_idx, val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN
            cell.font = BODY_FONT

    _autosize(ws, max_width=18)
    return header_row + len(firms) + 2


def build_workbook(
    companies: list[dict[str, Any]],
    commentary: list[dict[str, Any]],
    news: list[dict[str, Any]],
    peer_activity: list[dict[str, Any]],
    sector_calls: list[dict[str, Any]],
    meta: Optional[dict[str, Any]] = None,
    out_path: Optional[Path] = None,
) -> Path:
    meta = meta or {}
    out_path = Path(out_path) if out_path else DEFAULT_OUT
    out_path.parent.mkdir(parents=True, exist_ok=True)

    as_of = date.today()
    hot_cutoff_days = 30
    app_base = meta.get("app_base") or DEFAULT_APP_BASE
    alerts: list[dict[str, Any]] = list(meta.get("alerts") or [])
    digest: dict[str, Any] = dict(meta.get("digest") or {})
    thesis_shifts: list[dict[str, Any]] = list(meta.get("thesis_shifts") or [])

    wb = Workbook()

    commentary_by_co: dict[str, list[str]] = defaultdict(list)
    for cm in commentary:
        commentary_by_co[cm.get("company_id")].append(cm.get("quote_or_summary") or "")

    companies_sorted = sorted(companies, key=lambda x: -(x.get("thesis_score") or 0))
    deep = sum(1 for c in companies if c.get("recommendation") == "Deep Dive")
    watch = sum(1 for c in companies if c.get("recommendation") == "Watch")
    passed = sum(1 for c in companies if c.get("recommendation") == "Pass")
    stale_n = sum(1 for c in companies if c.get("is_stale"))
    hot_n = 0
    for c in companies:
        days = _days_since(c.get("last_signal_date"), as_of)
        if days is not None and days <= hot_cutoff_days and (
            c.get("recommendation") == "Deep Dive" or (c.get("thesis_score") or 0) >= 75
        ):
            hot_n += 1

    dominant = sum(1 for c in companies if c.get("pipeline_bucket") == "dominant_tech_growth")
    tactical = sum(
        1 for c in companies if c.get("pipeline_bucket") == "tactical_sector_agnostic"
    )
    n = max(1, len(companies))
    mix_dom = round(100 * dominant / n)
    mix_tac = round(100 * tactical / n)

    # ---------- Index (status only — no prose) ----------
    index = wb.active
    index.title = "Index"
    _set_tab_color(index, "Index")
    index["A1"] = "Thirdbase Signal — Deal Pipeline"
    index["A1"].font = TITLE_FONT
    index["A2"] = "Shared object of argument. Regenerated on every refresh / export."
    index["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    stats = [
        ("Refreshed", meta.get("last_refreshed", datetime.now(timezone.utc).isoformat())),
        ("Companies", len(companies)),
        ("Deep Dive / Watch / Pass", f"{deep} / {watch} / {passed}"),
        ("Hot Deals (30d)", hot_n),
        ("Stale (≥90d)", stale_n),
        ("Mix dominant/tactical %", f"{mix_dom} / {mix_tac} (target 60/40)"),
        ("Alerts", len(alerts)),
        ("Digest deals", len(digest.get("deals") or [])),
        ("File", str(out_path.name)),
        ("App", app_base),
    ]
    index["A4"] = "Status"
    index["A4"].font = LABEL_FONT
    for i, (k, v) in enumerate(stats, 5):
        index.cell(i, 1, k).font = BODY_FONT
        index.cell(i, 2, v).font = BODY_FONT

    index["A16"] = "Tabs"
    index["A16"].font = LABEL_FONT
    index["A17"] = "Core"
    index["B17"] = " · ".join(CORE_TABS)
    index["A18"] = "Signal extracts"
    index["B18"] = " · ".join(APP_TABS)
    index["A20"] = "Maintain"
    index["A20"].font = LABEL_FONT
    index["A21"] = "Full refresh"
    index["B21"] = "python scripts/refresh.py   OR header Refresh in Partner OS"
    index["A22"] = "Excel only"
    index["B22"] = "python scripts/export_workbook.py   OR POST /api/workbook"
    index["A23"] = "Download"
    index["B23"] = f"{app_base.rstrip('/')}/api/workbook"
    index.column_dimensions["A"].width = 26
    index.column_dimensions["B"].width = 88

    # ---------- Pipeline ----------
    ws = wb.create_sheet("Pipeline")
    _set_tab_color(ws, "Pipeline")
    pipe_rows: list[list[Any]] = []
    brief_urls: list[str] = []
    for c in companies_sorted:
        summary = c.get("commentary_summary") or "; ".join(
            commentary_by_co.get(c["id"], [])[:2]
        )
        url = _brief_url(c, app_base)
        brief_urls.append(url)
        bucket = c.get("pipeline_bucket") or ""
        if bucket == "dominant_tech_growth":
            bucket = "Dominant"
        elif bucket == "tactical_sector_agnostic":
            bucket = "Tactical"
        pipe_rows.append(
            [
                c.get("name"),
                _clip(c.get("one_liner"), 120),
                c.get("sector_theme"),
                c.get("stage"),
                bucket,
                c.get("last_round_size_m"),
                c.get("last_round_date"),
                c.get("valuation_est_m"),
                c.get("lead_investor"),
                c.get("tier1_count"),
                c.get("headcount"),
                c.get("headcount_6m_growth_pct"),
                c.get("thesis_score"),
                c.get("theme_id") or c.get("theme_tag"),
                c.get("last_signal_date"),
                c.get("recommendation"),
                "Open brief",
                _clip(summary, 180),
                _clip(c.get("why_now"), 140),
                c.get("relative_rank"),
            ]
        )
    _write_table(
        ws,
        PIPELINE_HEADERS,
        pipe_rows,
        widths={
            1: 18,
            2: 36,
            3: 22,
            17: 12,
            18: 40,
            19: 36,
            20: 14,
        },
        number_cols={6: "0.0", 8: "0.0", 12: "0.0", 13: "0.0"},
    )
    link_col = PIPELINE_HEADERS.index("Link to full brief") + 1
    for r_idx, url in enumerate(brief_urls, 2):
        cell = ws.cell(r_idx, link_col)
        cell.hyperlink = url
        cell.font = LINK_FONT
        cell.value = "Open brief"
    rec_col = PIPELINE_HEADERS.index("Recommendation") + 1
    score_col = PIPELINE_HEADERS.index("Thesis score") + 1
    _rec_formatting(ws, rec_col, len(pipe_rows))
    _score_scale(ws, score_col, len(pipe_rows))

    # ---------- Hot Deals ----------
    hot = wb.create_sheet("Hot Deals")
    _set_tab_color(hot, "Hot Deals")
    hot_rows: list[list[Any]] = []
    for c in companies_sorted:
        days = _days_since(c.get("last_signal_date"), as_of)
        if days is None or days > hot_cutoff_days:
            continue
        if not (
            c.get("recommendation") == "Deep Dive" or (c.get("thesis_score") or 0) >= 75
        ):
            continue
        hot_rows.append(
            [
                c.get("name"),
                _clip(c.get("one_liner"), 100),
                c.get("thesis_score"),
                c.get("recommendation"),
                c.get("sector_theme"),
                c.get("stage"),
                c.get("tier1_count"),
                c.get("lead_investor"),
                c.get("last_signal_date"),
                days,
                _clip(c.get("why_now"), 140),
                _brief_url(c, app_base),
            ]
        )
    _write_table(
        hot,
        [
            "Company",
            "One-line description",
            "Thesis score",
            "Recommendation",
            "Sector",
            "Stage",
            "Tier 1 count",
            "Lead investor",
            "Last signal date",
            "Days since signal",
            "Why now",
            "Link to full brief",
        ],
        hot_rows,
        widths={1: 18, 2: 34, 11: 40, 12: 12},
        number_cols={3: "0.0"},
    )
    _linkify(hot, 12, 2, len(hot_rows))
    _rec_formatting(hot, 4, len(hot_rows))
    _score_scale(hot, 3, len(hot_rows))

    # ---------- Watchlist ----------
    watch_ws = wb.create_sheet("Watchlist")
    _set_tab_color(watch_ws, "Watchlist")
    watch_rows = []
    for c in companies_sorted:
        if c.get("recommendation") == "Watch" or (c.get("stage") or "").lower() in (
            "seed",
            "pre-seed",
        ):
            watch_rows.append(
                [
                    c.get("name"),
                    _clip(c.get("one_liner"), 100),
                    c.get("stage"),
                    c.get("sector_theme"),
                    c.get("thesis_score"),
                    c.get("recommendation"),
                    c.get("last_signal_date"),
                    _days_since(c.get("last_signal_date"), as_of),
                    c.get("lead_investor"),
                    c.get("tier1_count"),
                    _brief_url(c, app_base),
                ]
            )
    _write_table(
        watch_ws,
        [
            "Company",
            "One-line description",
            "Stage",
            "Sector",
            "Thesis score",
            "Recommendation",
            "Last signal date",
            "Days silent",
            "Lead investor",
            "Tier 1 count",
            "Link to full brief",
        ],
        watch_rows,
        widths={1: 18, 2: 34, 11: 12},
        number_cols={5: "0.0"},
    )
    _linkify(watch_ws, 11, 2, len(watch_rows))
    _rec_formatting(watch_ws, 6, len(watch_rows))

    # ---------- Sector of Tomorrow ----------
    sot = wb.create_sheet("Sector of Tomorrow")
    _set_tab_color(sot, "Sector of Tomorrow")
    sot_rows = [
        [
            s.get("subsector"),
            s.get("parent_theme"),
            s.get("heat_score"),
            s.get("consensus_level"),
            _clip(" | ".join(s.get("evidence") or []), 160),
            ", ".join(s.get("top_companies") or []),
            _clip(s.get("why_thirdbase_cares"), 140),
        ]
        for s in sorted(sector_calls, key=lambda x: -(x.get("heat_score") or 0))
    ]
    _write_table(
        sot,
        [
            "Emerging sector",
            "Parent theme",
            "Heat score",
            "Consensus level",
            "Evidence",
            "Best companies found",
            "Why Thirdbase cares",
        ],
        sot_rows,
        widths={1: 32, 5: 40, 6: 28, 7: 36},
        number_cols={3: "0"},
    )
    _score_scale(sot, 3, len(sot_rows))

    # ---------- Peer Set Activity ----------
    peer = wb.create_sheet("Peer Set Activity")
    _set_tab_color(peer, "Peer Set Activity")
    matrix = meta.get("peer_matrix") or {}
    next_row = _write_peer_matrix(peer, matrix, start_row=1)
    if next_row == 1:
        peer.cell(1, 1, "Activity log — filter Theme / Stage / Date")
        peer.cell(1, 1).font = LABEL_FONT
        next_row = 2
    else:
        peer.cell(next_row, 1, "Activity log — filter Theme / Stage / Date")
        peer.cell(next_row, 1).font = LABEL_FONT
        next_row += 1

    company_by_name = {c.get("name"): c for c in companies}
    peer_rows = []
    for p in sorted(peer_activity, key=lambda x: str(x.get("date") or ""), reverse=True):
        co = company_by_name.get(p.get("company_name") or "")
        peer_rows.append(
            [
                p.get("firm"),
                p.get("company_name"),
                p.get("round") or (co.get("stage") if co else ""),
                p.get("date"),
                p.get("theme") or (co.get("sector_theme") if co else ""),
                co.get("stage") if co else (p.get("round") or ""),
                "Y" if p.get("on_thesis_flag") else "N",
                "Y" if p.get("thesis_shift") else "N",
                _clip(p.get("notes"), 120),
            ]
        )
    log_headers = [
        "Firm",
        "Company",
        "Round",
        "Date",
        "Theme",
        "Stage",
        "On thesis",
        "Thesis shift",
        "Notes",
    ]
    for i, h in enumerate(log_headers, 1):
        peer.cell(next_row, i, h)
        peer.cell(next_row, i).fill = HEADER_FILL
        peer.cell(next_row, i).font = HEADER_FONT
    for r_idx, row in enumerate(peer_rows, next_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = peer.cell(r_idx, c_idx, val)
            cell.border = THIN
            cell.font = BODY_FONT
            if c_idx == 8 and val == "Y":
                cell.fill = HIGH_FILL
    if peer_rows:
        peer.auto_filter.ref = (
            f"A{next_row}:{get_column_letter(len(log_headers))}{next_row + len(peer_rows)}"
        )
    _autosize(peer)

    # ---------- Co-investor Heatmap ----------
    heat = wb.create_sheet("Co-investor Heatmap")
    _set_tab_color(heat, "Co-investor Heatmap")
    heat_meta = meta.get("heatmap") or []
    if heat_meta:
        heat_rows = [
            [
                h.get("firm_a"),
                h.get("firm_b"),
                h.get("coinvest_count"),
                ", ".join(h.get("shared_themes") or []),
                h.get("last_shared_deal"),
                h.get("last_shared_date"),
                h.get("syndicate_score"),
            ]
            for h in heat_meta
        ]
        headers = [
            "Firm A",
            "Firm B",
            "Co-invest count",
            "Shared themes",
            "Last shared deal",
            "Last date",
            "Syndicate score",
        ]
    else:
        heat_rows = _coinvestor_pairs(companies)
        headers = [
            "Firm A",
            "Firm B",
            "Co-invest count",
            "Shared themes",
            "Last shared deal",
            "Last date",
        ]
    _write_table(heat, headers, heat_rows, widths={1: 20, 2: 20, 4: 36})

    # ---------- News Worth Reading ----------
    news_ws = wb.create_sheet("News Worth Reading")
    _set_tab_color(news_ws, "News Worth Reading")
    news_rows = []
    for n in news:
        kind = classify_news_kind(
            n.get("title") or "",
            n.get("source") or "",
            n.get("why_it_matters") or "",
        )
        news_rows.append(
            [
                _clip(n.get("title"), 100),
                n.get("source"),
                kind,
                n.get("published_at"),
                _clip(n.get("why_it_matters"), 140),
                ", ".join(n.get("related_themes") or []),
                n.get("url") or "",
            ]
        )
    _write_table(
        news_ws,
        [
            "Title",
            "Source",
            "Kind",
            "Published",
            "Why partners should read",
            "Related themes",
            "URL",
        ],
        news_rows,
        widths={1: 40, 5: 40, 7: 28},
    )
    for r_idx in range(2, len(news_rows) + 2):
        cell = news_ws.cell(r_idx, 7)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.font = LINK_FONT

    # ---------- Investor Commentary ----------
    com = wb.create_sheet("Investor Commentary")
    _set_tab_color(com, "Investor Commentary")
    com_rows = [
        [
            c.get("company_name"),
            c.get("source"),
            _clip(c.get("quote_or_summary"), 180),
            c.get("sentiment"),
            c.get("credibility_tier"),
            c.get("captured_at"),
        ]
        for c in commentary
    ]
    _write_table(
        com,
        ["Company", "Source", "Quote / summary", "Sentiment", "Credibility", "Captured at"],
        com_rows,
        widths={1: 16, 3: 48},
    )

    # ---------- Stale ----------
    stale = wb.create_sheet("Stale")
    _set_tab_color(stale, "Stale")
    stale_cos = [c for c in companies if c.get("is_stale")]
    stale_cos.sort(key=lambda c: _days_since(c.get("last_signal_date"), as_of) or 0, reverse=True)
    stale_rows = [
        [
            c.get("name"),
            c.get("sector_theme"),
            c.get("stage"),
            c.get("last_signal_date"),
            _days_since(c.get("last_signal_date"), as_of),
            c.get("thesis_score"),
            c.get("recommendation"),
            c.get("review_status") or "Pending Partner Review",
            _brief_url(c, app_base),
        ]
        for c in stale_cos
    ]
    _write_table(
        stale,
        [
            "Company",
            "Sector",
            "Stage",
            "Last signal date",
            "Days silent",
            "Thesis score",
            "Recommendation",
            "Review status",
            "Link to full brief",
        ],
        stale_rows,
        widths={1: 18, 8: 22, 9: 12},
        number_cols={6: "0.0"},
    )
    _linkify(stale, 9, 2, len(stale_rows))
    if stale_rows:
        stale.conditional_formatting.add(
            f"E2:E{len(stale_rows) + 1}",
            FormulaRule(formula=["E2>=120"], fill=HIGH_FILL),
        )

    # ---------- Alerts (Signal app) ----------
    alert_ws = wb.create_sheet("Alerts")
    _set_tab_color(alert_ws, "Alerts")
    severity_rank = {"high": 0, "medium": 1, "low": 2}
    alert_sorted = sorted(
        alerts,
        key=lambda a: (
            severity_rank.get(str(a.get("severity") or "").lower(), 9),
            str(a.get("created_at") or ""),
        ),
    )
    alert_rows = [
        [
            a.get("severity"),
            a.get("alert_type"),
            _clip(a.get("title"), 80),
            _clip(a.get("body"), 160),
            a.get("company_name")
            or next(
                (c.get("name") for c in companies if c.get("id") == a.get("company_id")),
                "",
            ),
            a.get("created_at"),
            (
                f"{app_base.rstrip('/')}{a.get('brief_url')}"
                if a.get("brief_url")
                else (
                    _brief_url(
                        next(
                            (c for c in companies if c.get("id") == a.get("company_id")),
                            {},
                        ),
                        app_base,
                    )
                    if a.get("company_id")
                    else ""
                )
            ),
        ]
        for a in alert_sorted
    ]
    _write_table(
        alert_ws,
        ["Severity", "Type", "Title", "Body", "Company", "Created", "Brief"],
        alert_rows,
        widths={1: 10, 2: 18, 3: 36, 4: 44, 7: 12},
    )
    for r_idx in range(2, len(alert_rows) + 2):
        sev = str(alert_ws.cell(r_idx, 1).value or "").lower()
        if sev == "high":
            alert_ws.cell(r_idx, 1).fill = HIGH_FILL
        elif sev == "medium":
            alert_ws.cell(r_idx, 1).fill = MED_FILL
        cell = alert_ws.cell(r_idx, 7)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.value = "Open"
            cell.font = LINK_FONT

    # ---------- Digest (latest M/W/F) ----------
    dig = wb.create_sheet("Digest")
    _set_tab_color(dig, "Digest")
    dig["A1"] = digest.get("subject") or "Latest M/W/F digest"
    dig["A1"].font = TITLE_FONT
    dig["A2"] = digest.get("generated_at") or ""
    dig["A2"].font = Font(name="Calibri", size=9, color="666666")

    deal_rows = [
        [
            d.get("name"),
            d.get("recommendation"),
            d.get("score") or d.get("thesis_score"),
            d.get("sector") or d.get("subsector"),
            d.get("stage"),
            d.get("lead_investor"),
            d.get("tier1_count"),
            _clip(d.get("rationale") or d.get("why_now") or d.get("one_liner"), 140),
            d.get("brief_url")
            or (
                f"{app_base.rstrip('/')}/company/{d.get('slug')}"
                if d.get("slug")
                else ""
            ),
        ]
        for d in (digest.get("deals") or [])
    ]
    next_r = _write_table(
        dig,
        [
            "Company",
            "Rec",
            "Score",
            "Sector",
            "Stage",
            "Lead",
            "Tier 1",
            "Why in digest",
            "Brief",
        ],
        deal_rows,
        start_row=4,
        widths={1: 18, 8: 44, 9: 12},
        number_cols={3: "0.0"},
    )
    for r_idx in range(5, 5 + len(deal_rows)):
        cell = dig.cell(r_idx, 9)
        if cell.value:
            href = str(cell.value)
            if href.startswith("/"):
                href = f"{app_base.rstrip('/')}{href}"
            cell.hyperlink = href
            cell.value = "Open"
            cell.font = LINK_FONT

    dig.cell(next_r, 1, "Sector calls in digest")
    dig.cell(next_r, 1).font = LABEL_FONT
    sec_rows = [
        [
            s.get("subsector"),
            s.get("consensus_level"),
            _clip(s.get("why") or s.get("why_thirdbase_cares"), 120),
            ", ".join(s.get("top_companies") or []),
        ]
        for s in (digest.get("sector_calls") or [])
    ]
    next_r = _write_table(
        dig,
        ["Sector", "Consensus", "Why", "Look at"],
        sec_rows,
        start_row=next_r + 1,
        widths={1: 28, 3: 40, 4: 28},
    )

    dig.cell(next_r, 1, "News in digest")
    dig.cell(next_r, 1).font = LABEL_FONT
    n_rows = [
        [
            n.get("title"),
            n.get("source"),
            n.get("kind"),
            _clip(n.get("why") or n.get("why_it_matters"), 120),
            n.get("url") or "",
        ]
        for n in (digest.get("news") or [])
    ]
    _write_table(
        dig,
        ["Title", "Source", "Kind", "Why", "URL"],
        n_rows,
        start_row=next_r + 1,
        widths={1: 40, 4: 40, 5: 28},
    )

    # ---------- Scores (dimension breakdown) ----------
    scores = wb.create_sheet("Scores")
    _set_tab_color(scores, "Scores")
    score_headers = [
        "Company",
        "Rec",
        "Total",
        "Relative rank",
        *[SCORE_DIM_LABELS[d] for d in SCORE_DIMS],
        "Brief",
    ]
    score_rows = []
    for c in companies_sorted:
        bd = c.get("score_breakdown") or {}
        score_rows.append(
            [
                c.get("name"),
                c.get("recommendation"),
                c.get("thesis_score"),
                c.get("relative_rank"),
                *[bd.get(d) for d in SCORE_DIMS],
                _brief_url(c, app_base),
            ]
        )
    _write_table(
        scores,
        score_headers,
        score_rows,
        widths={1: 18, 4: 14, len(score_headers): 12},
        number_cols={3: "0.0", **{i: "0.0" for i in range(5, 5 + len(SCORE_DIMS))}},
    )
    _linkify(scores, len(score_headers), 2, len(score_rows))
    _rec_formatting(scores, 2, len(score_rows))
    _score_scale(scores, 3, len(score_rows))
    for dim_i in range(5, 5 + len(SCORE_DIMS)):
        _score_scale(scores, dim_i, len(score_rows))

    # ---------- Thesis Shifts ----------
    shift_ws = wb.create_sheet("Thesis Shifts")
    _set_tab_color(shift_ws, "Thesis Shifts")
    if not thesis_shifts:
        # Fall back to peer_activity flagged shifts
        thesis_shifts = [p for p in peer_activity if p.get("thesis_shift")]
    shift_rows = [
        [
            s.get("firm"),
            s.get("company_name") or s.get("company"),
            s.get("theme") or s.get("round"),
            s.get("date"),
            _clip(s.get("notes") or s.get("insight"), 160),
            "Y" if s.get("thesis_shift", True) else "N",
        ]
        for s in thesis_shifts
    ]
    _write_table(
        shift_ws,
        ["Firm", "Company", "Theme / round", "Date", "Notes", "Shift"],
        shift_rows,
        widths={1: 20, 2: 18, 5: 48},
    )

    # ---------- Peer Firm Dossiers ----------
    firm_ws = wb.create_sheet("Peer Firm Dossiers")
    _set_tab_color(firm_ws, "Peer Firm Dossiers")
    firm_rows = []
    for f in meta.get("peer_firms") or []:
        themes = ", ".join(
            t.get("theme") if isinstance(t, dict) else str(t)
            for t in (f.get("top_themes") or [])[:3]
        )
        firm_rows.append(
            [
                f.get("name"),
                _clip(f.get("stated_focus"), 60),
                f.get("deal_count"),
                f.get("lead_count"),
                f.get("deep_dive_count"),
                f.get("drift_score"),
                f.get("focus_alignment"),
                f.get("conviction_score"),
                f.get("watch_priority"),
                f.get("thesis_shift_count"),
                themes,
                f.get("last_activity_date"),
                _clip(f.get("intel_summary"), 140),
            ]
        )
    firm_rows.sort(key=lambda r: -(r[8] or 0) if isinstance(r[8], (int, float)) else 0)
    _write_table(
        firm_ws,
        [
            "Firm",
            "Stated focus",
            "Deals",
            "Leads",
            "Deep Dives",
            "Drift",
            "Alignment",
            "Conviction",
            "Watch pri",
            "Shifts",
            "Top themes",
            "Last activity",
            "Intel",
        ],
        firm_rows,
        widths={1: 22, 2: 28, 11: 32, 13: 40},
    )

    # ---------- Actions (golden insights as queue) ----------
    act = wb.create_sheet("Actions")
    _set_tab_color(act, "Actions")
    brief = meta.get("golden_brief") or {}
    act["A1"] = "Partner action queue"
    act["A1"].font = TITLE_FONT
    must = brief.get("must_do") or []
    if must:
        act["A2"] = "Must do"
        act["A2"].font = LABEL_FONT
        act["B2"] = " · ".join(_clip(x, 80) for x in must[:5])
    prop = brief.get("proprietary") or []
    if prop:
        act["A3"] = "Proprietary windows"
        act["A3"].font = LABEL_FONT
        act["B3"] = ", ".join(str(x) for x in prop[:8])

    urgency_rank = {"now": 0, "this_week": 1, "monitor": 2}
    gold_rows = sorted(
        meta.get("golden_insights") or [],
        key=lambda g: (urgency_rank.get(str(g.get("urgency") or ""), 9), -(g.get("score") or 0)),
    )
    action_rows = [
        [
            g.get("urgency"),
            g.get("kind"),
            _clip(g.get("title"), 80),
            _clip(g.get("insight"), 140),
            _clip(g.get("action"), 120),
            g.get("score"),
        ]
        for g in gold_rows
    ]
    _write_table(
        act,
        ["Urgency", "Kind", "Title", "Insight", "Action", "Score"],
        action_rows,
        start_row=5,
        widths={1: 12, 2: 12, 3: 36, 4: 44, 5: 40},
    )
    for r_idx in range(6, 6 + len(action_rows)):
        urg = str(act.cell(r_idx, 1).value or "").lower()
        if urg == "now":
            act.cell(r_idx, 1).fill = HIGH_FILL
        elif urg == "this_week":
            act.cell(r_idx, 1).fill = MED_FILL

    # ---------- Judgment (tables only) ----------
    judgment = meta.get("judgment") or {}
    jws = wb.create_sheet("Judgment")
    _set_tab_color(jws, "Judgment")
    mix = judgment.get("mix") or {}
    jws["A1"] = "Judgment OS"
    jws["A1"].font = TITLE_FONT
    jws["A2"] = "Mix"
    jws["A2"].font = LABEL_FONT
    jws["B2"] = (
        f"{mix.get('dominant_pct', mix_dom)}/{mix.get('tactical_pct', mix_tac)} · "
        f"{mix.get('status') or '—'} · {_clip(mix.get('alarm') or mix.get('counsel') or '', 100)}"
    )

    jws["A4"] = "Miss retrospectives"
    jws["A4"].font = LABEL_FONT
    miss_rows = [
        [
            m.get("company"),
            m.get("then_rec"),
            m.get("then_score"),
            m.get("severity"),
            _clip(m.get("now_signal"), 100),
            _clip(m.get("lesson"), 100),
            _clip(m.get("action"), 100),
        ]
        for m in (judgment.get("misses") or [])
    ]
    next_r = _write_table(
        jws,
        ["Company", "Then rec", "Then score", "Severity", "Now signal", "Lesson", "Action"],
        miss_rows,
        start_row=5,
        widths={1: 16, 5: 28, 6: 28, 7: 28},
    )

    jws.cell(next_r, 1, "Founder radar")
    jws.cell(next_r, 1).font = LABEL_FONT
    fr_rows = [
        [
            f.get("founder"),
            f.get("prior"),
            f.get("urgency"),
            f.get("company"),
            _clip(f.get("signal"), 100),
            _clip(f.get("action"), 100),
            f.get("source"),
        ]
        for f in (judgment.get("founder_radar") or [])
    ]
    next_r = _write_table(
        jws,
        ["Founder", "Prior", "Urgency", "Company", "Signal", "Action", "Source"],
        fr_rows,
        start_row=next_r + 1,
        widths={1: 18, 5: 32, 6: 32},
    )

    jws.cell(next_r, 1, "Evidence freshness")
    jws.cell(next_r, 1).font = LABEL_FONT
    fx_rows = [
        [
            f.get("company"),
            f.get("overall"),
            f.get("score_confidence"),
            ", ".join(f.get("stale_fields") or [])
            if isinstance(f.get("stale_fields"), list)
            else f.get("stale_fields"),
        ]
        for f in (judgment.get("freshness") or [])
    ]
    _write_table(
        jws,
        ["Company", "Overall", "Score confidence", "Stale fields"],
        fx_rows,
        start_row=next_r + 1,
        widths={1: 18, 4: 40},
    )

    wb.save(out_path)
    return out_path
