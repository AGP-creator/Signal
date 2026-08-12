"""Partner watchlists — multi-partner ranked sets + Excel ingest."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.ingest.dedupe import find_duplicate, merge_company, normalize_name

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "data" / "output" / "Signal_Partner_Watchlist_Template.xlsx"
META_KEY = "partner_watchlists"

TEMPLATE_HEADERS = [
    "Company",
    "Domain",
    "Sector",
    "Stage",
    "Round size ($M)",
    "Round date",
    "Lead investor",
    "Investors",
    "One-liner",
    "Partner note",
    "Source",
]

HEADER_ALIASES = {
    "company": "name",
    "company name": "name",
    "name": "name",
    "domain": "domain",
    "website": "domain",
    "url": "domain",
    "sector": "sector_theme",
    "sector theme": "sector_theme",
    "theme": "sector_theme",
    "stage": "stage",
    "round size ($m)": "last_round_size_m",
    "round size": "last_round_size_m",
    "raise ($m)": "last_round_size_m",
    "round date": "last_round_date",
    "lead investor": "lead_investor",
    "lead": "lead_investor",
    "investors": "investors",
    "one-liner": "one_liner",
    "one liner": "one_liner",
    "description": "one_liner",
    "partner note": "note",
    "note": "note",
    "notes": "note",
    "source": "row_source",
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sid(prefix: str, *parts: str) -> str:
    h = hashlib.sha1("|".join(parts).encode()).hexdigest()[:10]
    return f"{prefix}_{h}"


def _slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s[:48] or "company"


def _cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _float_or_none(v: Any) -> Optional[float]:
    s = _cell(v).replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _split_investors(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[,;|/]+", raw)
    return [p.strip() for p in parts if p.strip()]


def build_template(out_path: Optional[Path] = None) -> Path:
    path = out_path or TEMPLATE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    fill = PatternFill("solid", fgColor="1B2A4A")
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    for i, h in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(1, i, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left")
    # Example row so partners see the shape
    example = [
        "Acme Agents",
        "acmeagents.com",
        "AI infra",
        "Seed",
        8,
        "2026-06-01",
        "Example Ventures",
        "Example Ventures, Angel X",
        "Agent runtime with eval harness for enterprise.",
        "Met at Demo Day — want a second look.",
        "Demo Day",
    ]
    for i, v in enumerate(example, 1):
        ws.cell(2, i, v)
    ws.auto_filter.ref = f"A1:K1"
    widths = [22, 20, 16, 10, 14, 12, 18, 28, 40, 32, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    # Fix column widths properly
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    guide = wb.create_sheet("Instructions", 0)
    guide["A1"] = "Signal — Partner Watchlist Template"
    guide["A1"].font = Font(bold=True, size=14, color="1B2A4A", name="Calibri")
    lines = [
        "1. Fill the Watchlist sheet (one company per row). Keep the header row.",
        "2. Company is required. Other columns are optional but improve matching.",
        "3. Upload in Interest Desk → Excel import. Preview matches before commit.",
        "4. Existing pipeline names are linked; new names enter as partner-sourced Watch.",
        "5. Each partner keeps their own ranked set; firm overlap shows who else added a name.",
    ]
    for i, line in enumerate(lines, 3):
        guide[f"A{i}"] = line
    guide.column_dimensions["A"].width = 96
    wb.save(path)
    return path


def _map_headers(raw_headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, h in enumerate(raw_headers):
        key = HEADER_ALIASES.get(_cell(h).lower())
        if key:
            mapping[idx] = key
    return mapping


def parse_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = None
    for name in ("Watchlist", "Sheet1", wb.sheetnames[0]):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_cell(h) for h in rows[0]]
    mapping = _map_headers(headers)
    if "name" not in mapping.values():
        # Try first column as company if headers are wrong
        mapping = {0: "name", **{k: v for k, v in mapping.items() if v != "name"}}
    out: list[dict[str, Any]] = []
    for r_i, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        item: dict[str, Any] = {"_row": r_i}
        for col_idx, field in mapping.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if field == "last_round_size_m":
                item[field] = _float_or_none(val)
            elif field == "investors":
                item[field] = _split_investors(_cell(val))
            else:
                item[field] = _cell(val) or None
        if not item.get("name"):
            continue
        out.append(item)
    return out


def skeleton_from_row(row: dict[str, Any], partner_name: str) -> dict[str, Any]:
    name = row["name"]
    cid = _sid("pw", normalize_name(name))
    domain = row.get("domain")
    if domain and "://" in domain:
        domain = domain.split("://", 1)[1].split("/")[0]
    investors = row.get("investors") or []
    if isinstance(investors, str):
        investors = _split_investors(investors)
    lead = row.get("lead_investor")
    if lead and lead not in investors:
        investors = [lead, *investors]
    return {
        "id": cid,
        "name": name,
        "slug": _slugify(name),
        "domain": domain,
        "one_liner": row.get("one_liner") or f"Partner-sourced by {partner_name}",
        "sector_theme": row.get("sector_theme") or "Unclassified",
        "theme_id": None,
        "subsector": row.get("sector_theme"),
        "stage": row.get("stage") or "Seed",
        "pipeline_bucket": "partner_sourced",
        "last_round_size_m": row.get("last_round_size_m"),
        "last_round_date": row.get("last_round_date"),
        "valuation_est_m": None,
        "valuation_confidence": "low",
        "lead_investor": lead,
        "investors": investors,
        "tier1_count": 0,
        "tier1_names": [],
        "tier2_count": 0,
        "headcount": None,
        "headcount_6m_growth_pct": None,
        "yoy_growth_pct": None,
        "runway_months_est": None,
        "tam_usd_b": None,
        "moat_notes": None,
        "team_notes": None,
        "traction_notes": None,
        "last_signal_date": (row.get("last_round_date") or _now()[:10]),
        "sources": ["partner_watchlist", row.get("row_source") or "excel"],
        "is_stale": False,
        "review_status": "Partner watchlist",
        "thesis_score": None,
        "score_breakdown": {},
        "relative_rank": None,
        "recommendation": "Watch",
        "why_now": row.get("note") or f"Added via partner watchlist ({partner_name})",
        "commentary_summary": None,
        "brief_id": f"brief_{cid}",
        "discovery_origin": {
            "source": "partner_excel",
            "partner": partner_name,
            "row_source": row.get("row_source"),
        },
        "update_events": [],
    }


def preview_rows(
    rows: list[dict[str, Any]],
    companies: list[dict[str, Any]],
    partner_name: str,
) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for row in rows:
        incoming = skeleton_from_row(row, partner_name)
        dup = find_duplicate(companies, incoming)
        if dup:
            action = "match"
            company_id = dup["id"]
            company_name = dup.get("name") or incoming["name"]
        else:
            action = "create"
            company_id = incoming["id"]
            company_name = incoming["name"]
        preview.append(
            {
                "row": row.get("_row"),
                "action": action,
                "company_id": company_id,
                "company_name": company_name,
                "name": row.get("name"),
                "domain": row.get("domain"),
                "sector_theme": row.get("sector_theme"),
                "stage": row.get("stage"),
                "note": row.get("note"),
                "row_source": row.get("row_source") or "excel",
                "payload": incoming if action == "create" else None,
                "merge_fields": {
                    k: row.get(k)
                    for k in (
                        "domain",
                        "sector_theme",
                        "stage",
                        "last_round_size_m",
                        "last_round_date",
                        "lead_investor",
                        "investors",
                        "one_liner",
                    )
                    if row.get(k)
                }
                if action == "match"
                else None,
            }
        )
    return preview


def load_watchlists_meta() -> dict[str, list[dict[str, Any]]]:
    from src.db.supabase_store import get_meta

    raw = get_meta(META_KEY, "{}")
    try:
        parsed = json.loads(raw or "{}")
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    out: dict[str, list[dict[str, Any]]] = {}
    for partner, items in parsed.items():
        if isinstance(items, list):
            out[str(partner)] = [i for i in items if isinstance(i, dict) and i.get("company_id")]
    return out


def save_watchlists_meta(data: dict[str, list[dict[str, Any]]]) -> None:
    from src.db.supabase_store import set_meta

    set_meta(META_KEY, json.dumps(data))


def load_watchlist_items() -> list[dict[str, Any]]:
    """Prefer table; fall back to meta."""
    from src.db.supabase_store import load_table

    try:
        rows = load_table("partner_watchlist_items")
        if rows:
            return rows
    except Exception:
        pass
    flat: list[dict[str, Any]] = []
    for partner, items in load_watchlists_meta().items():
        for it in items:
            flat.append({**it, "partner_name": partner})
    return flat


def upsert_watchlist_item(
    partner_name: str,
    company_id: str,
    *,
    rank: Optional[int] = None,
    note: Optional[str] = None,
    source: str = "ui",
) -> dict[str, Any]:
    from src.db.supabase_store import upsert_rows

    now = _now()
    meta = load_watchlists_meta()
    items = list(meta.get(partner_name) or [])
    existing = next((i for i in items if i.get("company_id") == company_id), None)
    if existing:
        if rank is not None:
            existing["rank"] = rank
        if note is not None:
            existing["note"] = note
        existing["source"] = source or existing.get("source") or "ui"
        existing["updated_at"] = now
        row = {**existing, "partner_name": partner_name}
    else:
        row = {
            "partner_name": partner_name,
            "company_id": company_id,
            "rank": rank if rank is not None else len(items) + 1,
            "note": note,
            "source": source or "ui",
            "added_at": now,
            "updated_at": now,
        }
        items.append({k: v for k, v in row.items() if k != "partner_name"})
    # re-normalize ranks
    items = sorted(items, key=lambda x: int(x.get("rank") or 999))
    for i, it in enumerate(items, 1):
        it["rank"] = i
        it["updated_at"] = it.get("updated_at") or now
    meta[partner_name] = items
    save_watchlists_meta(meta)

    table_row = next(
        (
            {**it, "partner_name": partner_name}
            for it in items
            if it.get("company_id") == company_id
        ),
        row,
    )
    try:
        upsert_rows("partner_watchlist_items", [table_row])
    except Exception:
        pass
    return table_row


def remove_watchlist_item(partner_name: str, company_id: str) -> None:
    from src.db.supabase_store import get_supabase

    meta = load_watchlists_meta()
    items = [i for i in (meta.get(partner_name) or []) if i.get("company_id") != company_id]
    for i, it in enumerate(items, 1):
        it["rank"] = i
    meta[partner_name] = items
    save_watchlists_meta(meta)
    try:
        get_supabase().table("partner_watchlist_items").delete().eq(
            "partner_name", partner_name
        ).eq("company_id", company_id).execute()
    except Exception:
        pass


def replace_partner_ranks(partner_name: str, ranked_ids: list[str]) -> list[dict[str, Any]]:
    meta = load_watchlists_meta()
    by_id = {i["company_id"]: i for i in (meta.get(partner_name) or []) if i.get("company_id")}
    now = _now()
    items: list[dict[str, Any]] = []
    for i, cid in enumerate(ranked_ids, 1):
        prev = by_id.get(cid) or {
            "company_id": cid,
            "note": None,
            "source": "ui",
            "added_at": now,
        }
        items.append(
            {
                **prev,
                "company_id": cid,
                "rank": i,
                "updated_at": now,
            }
        )
    meta[partner_name] = items
    save_watchlists_meta(meta)
    rows = [{**it, "partner_name": partner_name} for it in items]
    try:
        from src.db.supabase_store import get_supabase, upsert_rows

        # drop removed
        existing = [
            r
            for r in load_watchlist_items()
            if (r.get("partner_name") or "") == partner_name
        ]
        keep = set(ranked_ids)
        client = get_supabase()
        for r in existing:
            cid = r.get("company_id")
            if cid and cid not in keep:
                try:
                    client.table("partner_watchlist_items").delete().eq(
                        "partner_name", partner_name
                    ).eq("company_id", cid).execute()
                except Exception:
                    pass
        if rows:
            upsert_rows("partner_watchlist_items", rows)
    except Exception:
        pass
    return rows


def commit_import(
    preview: list[dict[str, Any]],
    companies: list[dict[str, Any]],
    partner_name: str,
    *,
    filename: str = "upload.xlsx",
) -> dict[str, Any]:
    from src.db.supabase_store import upsert_companies, upsert_rows

    working = list(companies)
    matched = created = skipped = 0
    errors: list[str] = []
    to_upsert: list[dict[str, Any]] = []
    watch_rows: list[dict[str, Any]] = []
    now = _now()

    # Start ranks after current list length
    meta = load_watchlists_meta()
    existing_items = list(meta.get(partner_name) or [])
    existing_ids = {i.get("company_id") for i in existing_items}
    next_rank = len(existing_items) + 1

    for item in preview:
        try:
            action = item.get("action")
            cid = item.get("company_id")
            if not cid:
                skipped += 1
                continue
            if action == "create":
                payload = item.get("payload") or skeleton_from_row(
                    {
                        "name": item.get("name") or item.get("company_name"),
                        "domain": item.get("domain"),
                        "sector_theme": item.get("sector_theme"),
                        "stage": item.get("stage"),
                        "note": item.get("note"),
                        "row_source": item.get("row_source"),
                    },
                    partner_name,
                )
                # re-check against working set
                dup = find_duplicate(working, payload)
                if dup:
                    cid = dup["id"]
                    fields = item.get("merge_fields") or {}
                    if fields:
                        merged = merge_company(dup, {**payload, **fields})
                        to_upsert.append(merged)
                        # replace in working
                        for i, c in enumerate(working):
                            if c.get("id") == dup["id"]:
                                working[i] = merged
                                break
                    matched += 1
                else:
                    to_upsert.append(payload)
                    working.append(payload)
                    cid = payload["id"]
                    created += 1
            else:
                fields = item.get("merge_fields") or {}
                if fields:
                    base = next((c for c in working if c.get("id") == cid), None)
                    if base:
                        patch = {
                            k: v
                            for k, v in {
                                "domain": fields.get("domain"),
                                "sector_theme": fields.get("sector_theme"),
                                "stage": fields.get("stage"),
                                "last_round_size_m": fields.get("last_round_size_m"),
                                "last_round_date": fields.get("last_round_date"),
                                "lead_investor": fields.get("lead_investor"),
                                "investors": fields.get("investors"),
                                "one_liner": fields.get("one_liner"),
                            }.items()
                            if v
                        }
                        if patch:
                            merged = merge_company(base, patch)
                            to_upsert.append(merged)
                            for i, c in enumerate(working):
                                if c.get("id") == cid:
                                    working[i] = merged
                                    break
                matched += 1

            if cid in existing_ids:
                # update note if provided
                for it in existing_items:
                    if it.get("company_id") == cid and item.get("note"):
                        it["note"] = item["note"]
                        it["updated_at"] = now
                        it["source"] = "excel"
                continue

            watch_rows.append(
                {
                    "partner_name": partner_name,
                    "company_id": cid,
                    "rank": next_rank,
                    "note": item.get("note"),
                    "source": "excel",
                    "added_at": now,
                    "updated_at": now,
                }
            )
            existing_items.append(
                {
                    "company_id": cid,
                    "rank": next_rank,
                    "note": item.get("note"),
                    "source": "excel",
                    "added_at": now,
                    "updated_at": now,
                }
            )
            existing_ids.add(cid)
            next_rank += 1
        except Exception as exc:
            errors.append(f"row {item.get('row')}: {exc}")
            skipped += 1

    if to_upsert:
        upsert_companies(to_upsert)

    # persist watchlist items
    existing_items = sorted(existing_items, key=lambda x: int(x.get("rank") or 999))
    for i, it in enumerate(existing_items, 1):
        it["rank"] = i
    meta[partner_name] = existing_items
    save_watchlists_meta(meta)
    if watch_rows:
        # rewrite ranks for new rows from meta
        by_id = {i["company_id"]: i for i in existing_items}
        synced = [
            {**by_id[r["company_id"]], "partner_name": partner_name}
            for r in watch_rows
            if r["company_id"] in by_id
        ]
        try:
            upsert_rows("partner_watchlist_items", synced)
        except Exception:
            pass

    import_id = _sid("imp", partner_name, filename, now)
    import_row = {
        "id": import_id,
        "partner_name": partner_name,
        "filename": filename,
        "status": "committed",
        "row_count": len(preview),
        "matched": matched,
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "created_at": now,
    }
    try:
        upsert_rows("partner_watchlist_imports", [import_row])
    except Exception:
        pass

    return {
        "ok": True,
        "import": import_row,
        "matched": matched,
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "added_to_list": len(watch_rows),
    }
